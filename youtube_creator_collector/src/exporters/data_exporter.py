"""
数据导出模块
============

将分析结果导出为多种格式。
"""

import csv
import json
from pathlib import Path
from typing import List, Dict, Optional, Any
from datetime import datetime
import logging
from abc import ABC, abstractmethod

from config.settings import Settings
from models.creator import YouTubeCreator

class BaseExporter(ABC):
    """
    导出器基类
    """
    
    def __init__(self, settings: Settings):
        """
        初始化导出器
        
        Args:
            settings: 配置对象
        """
        self.settings = settings
        
    @abstractmethod
    def export(self, 
              creators: List[YouTubeCreator], 
              output_path: Path,
              options: Dict = None) -> bool:
        """
        导出数据
        
        Args:
            creators: 创作者列表
            output_path: 输出路径
            options: 导出选项
            
        Returns:
            是否成功
        """
        pass
    
    @abstractmethod
    def get_extension(self) -> str:
        """
        获取文件扩展名
        
        Returns:
            扩展名
        """
        pass

class CSVExporter(BaseExporter):
    """
    CSV格式导出器
    """
    
    def __init__(self, settings: Settings):
        super().__init__(settings)
        self.delimiter = settings.CSV_DELIMITER
        
    def export(self,
              creators: List[YouTubeCreator],
              output_path: Path,
              options: Dict = None) -> bool:
        """
        导出为CSV格式
        
        Args:
            creators: 创作者列表
            output_path: 输出路径
            options: 导出选项
            
        Returns:
            是否成功
        """
        if options is None:
            options = {}
            
        try:
            with open(output_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.DictWriter(f, 
                                       fieldnames=self._get_fieldnames(),
                                       delimiter=self.delimiter)
                
                writer.writeheader()
                
                for creator in creators:
                    row = self._creator_to_row(creator, options)
                    writer.writerow(row)
                    
            logging.info(f"Exported {len(creators)} creators to {output_path}")
            return True
            
        except Exception as e:
            logging.error(f"Failed to export CSV: {e}")
            return False
    
    def get_extension(self) -> str:
        return 'csv'
    
    def _get_fieldnames(self) -> List[str]:
        """获取CSV字段名"""
        return [
            'channel_id',
            'channel_name',
            'custom_url',
            'description',
            'subscriber_count',
            'total_views',
            'video_count',
            'view_per_subscriber',
            'engagement_rate',
            'subscriber_growth_rate',
            'view_growth_rate',
            'target_region_view_percentage',
            'email',
            'business_email',
            'twitter',
            'website',
            'last_video_date',
            'videos_last_30_days',
            'upload_frequency_per_week',
            'is_pc_game_focus',
            'status',
            'region_view_percentages',
            'tags',
            'last_updated',
            'collected_at'
        ]
    
    def _creator_to_row(self, creator: YouTubeCreator, options: Dict) -> Dict:
        """
        将创作者转换为CSV行
        
        Args:
            creator: 创作者对象
            options:        Returns:
            导出选项
            
 行数据字典
        """
        row = {
            'channel_id': creator.channel_id,
            'channel_name': creator.channel_name,
            'custom_url': creator.custom_url or '',
            'description': creator.description[:500] if creator.description else '',
        }
        
        if creator.statistics:
            row['subscriber_count'] = creator.statistics.subscriber_count
            row['total_views'] = creator.statistics.total_views
            row['video_count'] = creator.statistics.video_count
            row['view_per_subscriber'] = round(creator.statistics.view_per_subscriber, 2)
            row['engagement_rate'] = round(creator.statistics.engagement_rate, 2)
            row['subscriber_growth_rate'] = round(creator.statistics.subscriber_growth_rate, 4)
            row['view_growth_rate'] = round(creator.statistics.view_growth_rate, 4)
        else:
            row['subscriber_count'] = ''
            row['total_views'] = ''
            row['video_count'] = ''
            row['view_per_subscriber'] = ''
            row['engagement_rate'] = ''
            row['subscriber_growth_rate'] = ''
            row['view_growth_rate'] = ''
            
        # 目标地区观看占比
        target_regions = options.get('target_regions', self.settings.TARGET_REGIONS)
        row['target_region_view_percentage'] = round(
            creator.get_target_region_view_percentage(target_regions) * 100, 2
        )
        
        # 联系信息
        if creator.contact_info:
            row['email'] = creator.contact_info.email or ''
            row['business_email'] = creator.contact_info.business_email or ''
            row['twitter'] = creator.contact_info.twitter or ''
            row['website'] = creator.contact_info.website or ''
        else:
            row['email'] = ''
            row['business_email'] = ''
            row['twitter'] = ''
            row['website'] = ''
            
        # 活跃度数据
        if creator.activity_data:
            row['last_video_date'] = creator.activity_data.last_video_date.isoformat() if creator.activity_data.last_video_date else ''
            row['videos_last_30_days'] = creator.activity_data.total_videos_last_30_days
            row['upload_frequency_per_week'] = round(creator.activity_data.upload_frequency_per_week, 2)
        else:
            row['last_video_date'] = ''
            row['videos_last_30_days'] = ''
            row['upload_frequency_per_week'] = ''
            
        # 其他字段
        row['is_pc_game_focus'] = 'Yes' if creator.pc_game_focus else 'No'
        row['status'] = creator.status.value if creator.status else ''
        
        # 地区数据
        region_percentages = {}
        for region in creator.region_data:
            region_percentages[region.region_code] = region.view_percentage
        row['region_view_percentages'] = json.dumps(region_percentages)
        
        row['tags'] = ', '.join(creator.tags)
        row['last_updated'] = creator.last_updated.isoformat()
        row['collected_at'] = creator.collected_at.isoformat()
        
        return row

class JSONExporter(BaseExporter):
    """
    JSON格式导出器
    """
    
    def export(self,
              creators: List[YouTubeCreator],
              output_path: Path,
              options: Dict = None) -> bool:
        """
        导出为JSON格式
        
        Args:
            creators: 创作者列表
            output_path: 输出路径
            options: 导出选项
            
        Returns:
            是否成功
        """
        if options is None:
            options = {}
            
        try:
            data = {
                'exported_at': datetime.now().isoformat(),
                'total_creators': len(creators),
                'target_regions': options.get('target_regions', self.settings.TARGET_REGIONS),
                'creators': [creator.to_dict() for creator in creators]
            }
            
            indent = options.get('indent', self.settings.JSON_INDENT)
            
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(data, f, indent=indent, ensure_ascii=False)
                
            logging.info(f"Exported {len(creators)} creators to {output_path}")
            return True
            
        except Exception as e:
            logging.error(f"Failed to export JSON: {e}")
            return False
    
    def get_extension(self) -> str:
        return 'json'

class ExcelExporter(BaseExporter):
    """
    Excel格式导出器
    """
    
    def export(self,
              creators: List[YouTubeCreator],
              output_path: Path,
              options: Dict = None) -> bool:
        """
        导出为Excel格式
        
        Args:
            creators: 创作者列表
            output_path: 输出路径
            options: 导出选项
            
        Returns:
            是否成功
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            from openpyxl.utils.dataframe import dataframe_to_rows
        except ImportError:
            logging.error("openpyxl not installed. Install with: pip install openpyxl")
            return False
            
        if options is None:
            options = {}
            
        try:
            wb = openpyxl.Workbook()
            
            # 移除默认工作表
            default_sheet = wb.active
            wb.remove(default_sheet)
            
            # 创建主数据表
            ws = wb.create_sheet("Creators")
            
            # 设置表头样式
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            
            # 写入数据
            fieldnames = [
                'Channel ID', 'Channel Name', 'URL', 'Subscribers', 'Total Views',
                'Videos', 'Engagement Rate', 'Target Region %', 'Email', 'Last Video',
                'Videos/30 Days', 'Status', 'PC Game Focus', 'Tags'
            ]
            
            for col, header in enumerate(fieldnames, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
                
            # 写入行数据
            target_regions = options.get('target_regions', self.settings.TARGET_REGIONS)
            
            for row_idx, creator in enumerate(creators, 2):
                ws.cell(row=row_idx, column=1, value=creator.channel_id)
                ws.cell(row=row_idx, column=2, value=creator.channel_name)
                ws.cell(row=row_idx, column=3, value=creator.custom_url or '')
                
                if creator.statistics:
                    ws.cell(row=row_idx, column=4, value=creator.statistics.subscriber_count)
                    ws.cell(row=row_idx, column=5, value=creator.statistics.total_views)
                    ws.cell(row=row_idx, column=6, value=creator.statistics.video_count)
                    ws.cell(row=row_idx, column=7, value=round(creator.statistics.engagement_rate, 2))
                else:
                    ws.cell(row=row_idx, column=4, value=0)
                    ws.cell(row=row_idx, column=5, value=0)
                    ws.cell(row=row_idx, column=6, value=0)
                    ws.cell(row=row_idx, column=7, value=0)
                    
                region_pct = creator.get_target_region_view_percentage(target_regions) * 100
                ws.cell(row=row_idx, column=8, value=round(region_pct, 2))
                
                if creator.contact_info:
                    email = creator.contact_info.email or creator.contact_info.business_email or ''
                    ws.cell(row=row_idx, column=9, value=email)
                else:
                    ws.cell(row=row_idx, column=9, value='')
                    
                if creator.activity_data:
                    last_video = creator.activity_data.last_video_date
                    ws.cell(row=row_idx, column=10, 
                           value=last_video.isoformat() if last_video else '')
                    ws.cell(row=row_idx, column=11, 
                           value=creator.activity_data.total_videos_last_30_days)
                else:
                    ws.cell(row=row_idx, column=10, value='')
                    ws.cell(row=row_idx, column=11, value=0)
                    
                ws.cell(row=row_idx, column=12, value=creator.status.value if creator.status else '')
                ws.cell(row=row_idx, column=13, value='Yes' if creator.pc_game_focus else 'No')
                ws.cell(row=row_idx, column=14, value=', '.join(creator.tags))
                
            # 调整列宽
            for col in range(1, 15):
                ws.column_dimensions[chr(64 + col)].width = 20
                
            # 创建统计摘要表
            ws_summary = wb.create_sheet("Summary")
            
            ws_summary.cell(row=1, column=1, value="Metric")
            ws_summary.cell(row=1, column=2, value="Value")
            
            summary_data = [
                ("Total Creators", len(creators)),
                ("Total Subscribers", sum(c.statistics.subscriber_count for c in creators if c.statistics)),
                ("Total Views", sum(c.statistics.total_views for c in creators if c.statistics)),
                ("Average Engagement", round(mean(c.statistics.engagement_rate for c in creators if c.statistics), 2) if creators else 0),
                ("Export Date", datetime.now().isoformat())
            ]
            
            for row_idx, (metric, value) in enumerate(summary_data, 2):
                ws_summary.cell(row=row_idx, column=1, value=metric)
                ws_summary.cell(row=row_idx, column=2, value=value)
                
            # 保存文件
            wb.save(output_path)
            
            logging.info(f"Exported {len(creators)} creators to {output_path}")
            return True
            
        except Exception as e:
            logging.error(f"Failed to export Excel: {e}")
            return False
    
    def get_extension(self) -> str:
        return 'xlsx'

class DataExporter:
    """
    数据导出管理器
    
    协调多种格式的导出
    """
    
    EXPORTERS = {
        'csv': CSVExporter,
        'json': JSONExporter,
        'excel': ExcelExporter
    }
    
    def __init__(self, settings: Settings):
        """
        初始化导出器
        
        Args:
            settings: 配置对象
        """
        self.settings = settings
        self.logger = logging.getLogger(__name__)
        
    def export_results(self,
                      creators: List[YouTubeCreator],
                      formats: List[str] = None,
                      output_dir: Path = None,
                      options: Dict = None) -> Dict[str, Path]:
        """
        批量导出结果
        
        Args:
            creators: 创作者列表
            formats: 导出格式列表
            output_dir: 输出目录
            options: 导出选项
            
        Returns:
            导出文件路径字典
        """
        if formats is None:
            formats = self.settings.EXPORT_FORMATS
            
        if output_dir is None:
            output_dir = self.settings.OUTPUT_DIR
            
        if options is None:
            options = {}
            
        output_dir.mkdir(parents=True, exist_ok=True)
        
        results = {}
        
        for fmt in formats:
            if fmt not in self.EXPORTERS:
                self.logger.warning(f"Unknown format: {fmt}")
                continue
                
            exporter = self.EXPORTERS[fmt](self.settings)
            
            # 生成文件名
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"youtube_creators_{timestamp}.{exporter.get_extension()}"
            output_path = output_dir / filename
            
            # 执行导出
            if exporter.export(creators, output_path, options):
                results[fmt] = output_path
                
        return results
    
    def export_with_analysis(self,
                            creators: List[YouTubeCreator],
                            analysis_results: List,
                            formats: List[str] = None,
                            output_dir: Path = None) -> Dict[str, Path]:
        """
        导出包含分析结果的数据
        
        Args:
            creators: 创作者列表
            analysis_results: 分析结果列表
            formats: 导出格式列表
            output_dir: 输出目录
            
        Returns:
            导出文件路径字典
        """
        if formats is None:
            formats = ['csv', 'json']
            
        if output_dir is None:
            output_dir = self.settings.OUTPUT_DIR
            
        output_dir.mkdir(parents=True, exist_ok=True)
        
        results = {}
        
        # 导出主数据
        results.update(self.export_results(creators, formats, output_dir))
        
        # 导出分析报告
        if analysis_results:
            report_data = {
                'exported_at': datetime.now().isoformat(),
                'total_analyzed': len(analysis_results),
                'qualified': sum(1 for r in analysis_results if r.metrics.get('meets_all_criteria', False)),
                'results': [
                    {
                        'channel_id': r.creator_id,
                        'channel_name': r.creator_name,
                        'scores': r.scores,
                        'metrics': r.metrics,
                        'flags': r.flags,
                        'recommendations': r.recommendations,
                        'analyzed_at': r.analyzed_at.isoformat()
                    }
                    for r in analysis_results
                ]
            }
            
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            report_path = output_dir / f"analysis_report_{timestamp}.json"
            
            try:
                with open(report_path, 'w', encoding='utf-8') as f:
                    json.dump(report_data, f, indent=2, ensure_ascii=False)
                    
                results['analysis_report'] = report_path
                
            except Exception as e:
                self.logger.error(f"Failed to export analysis report: {e}")
                
        return results
